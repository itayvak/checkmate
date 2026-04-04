"""Export project comment library to the standard shared .xlsx layout (Hebrew column headers)."""

from __future__ import annotations

import io
import re
from typing import Any

# Matches the conventional "implementation" category used in shared grading spreadsheets.
DEFAULT_EXPORT_TYPE = "מימוש"

HEADERS = ("סוג", "הערה", "סנכרון", "תוספת", "תמחור", "תמחור מקסימלי")


def _safe_sheet_title(name: str) -> str:
    """Excel worksheet name: max 31 chars, no : \\ / ? * [ ]."""
    t = re.sub(r'[:\\/\?\*\[\]]', "_", (name or "").strip()) or "comments"
    return t[:31]


def comment_library_export_filename(project_name: str) -> str:
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", (project_name or "").strip()) or "comment_library"
    return f"{base[:120]}_comments.xlsx"


def build_comment_library_xlsx_bytes(
    comments: list[dict[str, Any]],
    *,
    sheet_title: str,
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _safe_sheet_title(sheet_title)

    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for c in comments:
        if not isinstance(c, dict):
            continue
        title = str(c.get("title") or "").strip()
        details = str(c.get("details") or "").strip()
        teacher = str(c.get("teacher_text") or "").strip()
        try:
            points = int(c.get("points") or 0)
        except (TypeError, ValueError):
            points = 0
        try:
            max_pts = int(c.get("max_points") if c.get("max_points") is not None else 100)
        except (TypeError, ValueError):
            max_pts = 100

        ws.cell(row=row, column=1, value=DEFAULT_EXPORT_TYPE)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=teacher)
        ws.cell(row=row, column=4, value=details)
        ws.cell(row=row, column=5, value=points)
        ws.cell(row=row, column=6, value=max_pts)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
