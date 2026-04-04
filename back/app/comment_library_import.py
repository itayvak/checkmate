"""Parse shared comment-library .xlsx files (same layout as export)."""

from __future__ import annotations

import io
from typing import Any

from .comment_library_export import HEADERS


def _data_start_row(ws: Any) -> int:
    """Row 1 is a header row when column B matches the standard הערה header."""
    b1 = ws.cell(row=1, column=2).value
    if b1 is not None and str(b1).strip() == HEADERS[1]:
        return 2
    return 1


def _as_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return max(0, value)
    if isinstance(value, float):
        return max(0, int(round(value)))
    s = str(value).strip().replace(",", ".")
    if not s:
        return default
    try:
        return max(0, int(float(s)))
    except (TypeError, ValueError):
        return default


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def parse_comment_library_xlsx(data: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Returns (rows_to_import, skipped).
    Each row: title, details, teacher_text, points, max_points, sheet_row.
    Each skipped: sheet_row, reason.
    """
    from openpyxl import load_workbook

    if not data:
        return [], [{"sheet_row": 0, "reason": "empty file"}]

    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        return [], [{"sheet_row": 0, "reason": "no worksheet"}]

    start = _data_start_row(ws)
    rows_out: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    def row_has_content(row_idx: int) -> bool:
        for c in range(1, 7):
            v = ws.cell(row=row_idx, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return True
        return False

    for r in range(start, ws.max_row + 1):
        title = _as_str(ws.cell(row=r, column=2).value)
        teacher_text = _as_str(ws.cell(row=r, column=3).value)
        details = _as_str(ws.cell(row=r, column=4).value)
        points = _as_int(ws.cell(row=r, column=5).value, 0)
        max_points = _as_int(ws.cell(row=r, column=6).value, 100)

        if not title:
            if not row_has_content(r):
                continue
            skipped.append({"sheet_row": r, "reason": "empty title"})
            continue

        rows_out.append(
            {
                "title": title,
                "details": details,
                "teacher_text": teacher_text,
                "points": points,
                "max_points": max_points,
                "sheet_row": r,
            }
        )

    return rows_out, skipped
